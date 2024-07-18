import json
from openpyxl import Workbook, load_workbook
class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb['Sheet1']

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)
    
    def get_json_all(self):
        headers = [cell.value for cell in next(self.ws.iter_rows(min_row=1, max_row=1))]
        characters = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            character = dict(zip(headers, row))
            characters.append(character)
        json_str = json.dumps(characters, ensure_ascii=False, indent=4)
        print(json_str)
        return json_str
    
    def get_json_by_row(self, row_number):
        row_data = []
        for cell in self.ws[row_number]:
            row_data.append(cell.value)
        headers = []
        for cell in self.ws[1]:
            headers.append(cell.value)
        row_dict = {headers[i]: row_data[i] for i in range(len(headers))}
        json_str = json.dumps(row_dict, ensure_ascii=False)
        print(json_str)
        return json_str
    
'''    
characters_path = './剧本信息/角色设定/characters.xlsx'
a = ExcelOp(characters_path)
a.get_json_by_row(3)'''