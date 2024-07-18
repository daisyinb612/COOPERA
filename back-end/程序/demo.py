from zhipuai import ZhipuAI
from langchain_community.document_loaders import PyPDFLoader,TextLoader,UnstructuredWordDocumentLoader
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.embeddings import HuggingFaceBgeEmbeddings
from langchain.vectorstores import Chroma
from llama_index.postprocessor.flag_embedding_reranker import FlagEmbeddingReranker
from llama_index.core.schema import NodeWithScore, QueryBundle, TextNode
#from langchain_community.document_loaders import DirectoryLoader
from openpyxl import Workbook, load_workbook
import os
import re

class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb['Sheet1']

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

class Asker(object):
    def __init__(self,apikey=None,knowledge=None):
        self.setting_role_create = '''
        假设你是一位剧作家，
        你的任务是根据我在接下来的对话中给出的###世界观设定###，设计出多个故事的主要角色，输出时参照###输出示例###
        ###输出示例###
        <角色A>：{姓名：***；身份：***；性格：***；经历：***；目标：***}
        <角色B>：{姓名：***；身份：***；性格：***；经历：***；目标：***}
        <角色C>：{姓名：***；身份：***；性格：***；经历：***；目标：***}
        <角色D>：{姓名：***；身份：***；性格：***；经历：***；目标：***}
        <角色E>：{姓名：***；身份：***；性格：***；经历：***；目标：***}
        '''
        self.setting_outline_create = '''
        假设你是一位剧作家，
        你的任务是根据我在接下来的对话中给出的###世界观设定###和###角色设定###，编写故事大纲，输出时参照###输出示例###
        ###输出示例###
        <章节1>：{章节名：***；情节梗概：***}
        <章节2>：{章节名：***；情节梗概：***}
        <章节3>：{章节名：***；情节梗概：***}
        <章节4>：{章节名：***；情节梗概：***}
        <章节5>：{章节名：***；情节梗概：***}
        '''
        self.setting_detail_create = ''''''
        self.apikey = apikey
        self.knowledgeid = knowledge
        self.history = None

    def save_history(self,question,answer,prompt=None):
        if self.history == None:
            self.history = [{"role": "system", "content": prompt}]
        self.history.append({"role": "user", "content": question})
        self.history.append({"role": "assistant", "content": answer})
        
    def ask(self,question,apikey=None,Knowledge=None, history=None):
        if history is None:
            client = ZhipuAI(api_key=self.apikey)
            response = client.chat.completions.create(
                model="glm-4",
                messages=[
                    {"role": "system", "content": question},
                    {"role": "user", "content": question},
                ],
                top_p=0.7,
                temperature=0.95,
                max_tokens=8192,
                stream=True,
            )
        else:
            client = ZhipuAI(api_key=self.apikey)
            response = client.chat.completions.create(
                model="glm-4",
                messages=history,
                top_p=0.7,
                temperature=0.95,
                max_tokens=8192,
                stream=True,
            )
        answer = ''
        print('思考中', end='')
        for trunk in response:
            print(trunk.choices[0].delta.content, end='')
            answer += trunk.choices[0].delta.content
        print('\n')
        # print(answer)
        # print('\n')
        print('提问完成')
        print('\n')
        return answer
    
    def role_answer_analyze(self,answer,filepath="D://大学//大二下//剧本生成//剧本信息//角色设定//角色表.xlsx"):
        pattern = r'<[^>]+>+：'
        roles = re.split(pattern,answer)
        #print(roles)
        file = ExcelOp(file=filepath)
        for i in range(1,len(roles)):
            detail = (roles[i].split("；"))
            for j in range(0,len(detail)):
                #detail[j].replace('{','')
                #detail[j].replace('}','')
                file.set_cell_value(i,j+1,detail[j])

    def outline_answer_analyze(self,answer,filepath="D://大学//大二下//剧本生成//剧本信息//故事大纲//大纲表.xlsx"):
        pattern = r'<[^>]+>+：'
        roles = re.split(pattern,answer)
        #print(roles)
        file = ExcelOp(file=filepath)
        for i in range(1,len(roles)):
            detail = (roles[i].split("；"))
            for j in range(0,len(detail)):
                #detail[j].replace('{','')
                #detail[j].replace('}','')
                file.set_cell_value(i,j+1,detail[j])

    def detail_answer_analyze(self,answer):
        pass

class RAG(object):
    def __init__(self):
        self.world_path = "D://大学//大二下//剧本生成//剧本信息//世界观设定"
        self.role_path = "D://大学//大二下//剧本生成//剧本信息//角色设定"
        self.outline_path = "D://大学//大二下//剧本生成//剧本信息//故事大纲"
        self.detail_path = "D://大学//大二下//剧本生成//剧本信息//具体情节"
        self.vectordb = None

    def save_file(self,path,persist_directory = "D://大学//大二下//剧本生成//剧本信息//向量数据库"):
        files = [os.path.join(path, file) for file in os.listdir(path)]
        data = []
        for file in files:
            print(file)
            if ".txt" in file:   
                loader = TextLoader(file,encoding='utf-8')
            if ".pdf" in file:   
                loader = PyPDFLoader(file)
            if ".docx" in file:
                loader = UnstructuredWordDocumentLoader(file)
            pages = loader.load_and_split()
            text_splitter = CharacterTextSplitter(
                separator="\n",
                chunk_size=50,
                chunk_overlap=20,
                length_function=len
            )
            text = text_splitter.split_documents(pages)
            embeddings = HuggingFaceBgeEmbeddings(model_name="BAAI/bge-small-zh-v1.5")
            vectordb = Chroma.from_documents(
                documents=text,
                embedding=embeddings,
                persist_directory=persist_directory
            )
            #print("成功转换为向量数据库数据的条目数："+vectordb._collection.count())
            vectordb.persist()
            self.vectordb = vectordb
            
    def search_from_db(self,question,persist_directory="D://大学//大二下//剧本生成//剧本信息//向量数据库"):
        vectorstore = self.vectordb
        search_result = vectorstore.similarity_search(question, k=3)
        reranker_text = []
        for i in range(3):
            reranker_text.append(search_result[i].page_content)
        reranker = FlagEmbeddingReranker(
            top_n=3,
            model="BAAI/bge-reranker-large",
            use_fp16=False
        )
        nodes = [NodeWithScore(node=TextNode(text=doc)) for doc in reranker_text]
        query_bundle = QueryBundle(query_str=question)
        ranked_nodes = reranker._postprocess_nodes(nodes, query_bundle)
        for node in ranked_nodes:
            print(node.node.get_content(), "-> Score:", node.score)
        #print(ranked_nodes)
        def get_score(item):
            return item.score
        ranked_nodes.sort(key=get_score, reverse=True)
        #print(ranked_nodes)
        print("最高相关的语料为："+ranked_nodes[0].node.get_content())
        return ranked_nodes[0].node.get_content()
    
    def get_world_setting(self,filepath):
        with open(filepath,"r",encoding="utf-8") as f:
            data = f.read()
            print(data)
            return "/n"+"###世界观设定###/n"+data
    
    def get_role_setting(self,filepath):
        excel = ExcelOp(file=filepath)
        result = "/n"+"###世界观设定###/n"
        for i in range(1,excel.get_row_clo_num()[0]):
            for j in range(1,excel.get_row_clo_num()[1]):
                result += excel.get_cell_value(i,j)
            result += '\n'
        return result
        
def role_create_test():
    api_key = "3e1947fcfd64a1c07ad2cfd0c171cf73.cukGjaNSuP8lItF5"
    #a = RAG()
    #a.save_file(a.world_path)
    #a.search_from_db("地球遭受了一场名为什么的全球性灾难？")
    a = Asker()
    r = RAG()
    a.apikey = api_key
    world_setting = r.get_world_setting("D://大学//大二下//剧本生成//剧本信息//世界观设定//1.txt")
    answer = a.ask(a.setting_role_create+world_setting)
    a.role_answer_analyze(answer=answer)

def outline_creat_test():
    api_key = "3e1947fcfd64a1c07ad2cfd0c171cf73.cukGjaNSuP8lItF5"
    a = Asker()
    r = RAG()
    a.apikey = api_key
    world_setting = r.get_world_setting("D://大学//大二下//剧本生成//剧本信息//世界观设定//1.txt")
    role_setting = r.get_role_setting("D://大学//大二下//剧本生成//剧本信息//角色设定//角色表.xlsx")
    question = a.setting_outline_create+world_setting+role_setting
    answer = a.ask(question)
    a.outline_answer_analyze(answer)

def RAG_test():
    r = RAG()
    r.save_file(r.world_path)
    result = r.search_from_db("地球遭受了一场名为什么的全球性灾难？")
    return result

if __name__ == '__main__':
    role_create_test()
    RAG_test()
    outline_creat_test()